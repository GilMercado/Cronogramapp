import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from datetime import date

st.set_page_config(page_title="Actualizar Cronograma", layout="centered")
st.title("🎬 Actualizador de Cronograma de Producción")

# Ruta del archivo original
ruta_excel = "tres de AcademiaOnSet_CHIPOCLES_Cronograma de producción.xlsx"

# Cargar el archivo Excel sin modificar su formato
wb = load_workbook(ruta_excel)
ws = wb.active

st.markdown("Modifica la información del proyecto sin alterar el formato del archivo original.")

# Entradas del usuario
nombre_proyecto = st.text_input("Nombre del Proyecto", value="Nuevo Proyecto")
fecha_actualizacion = st.date_input("Fecha de Actualización", date.today())

# Fases: editar solo las fechas sin alterar estructura (filas 5 a 18 en Excel 1-indexed)
inicio_filas = 5
fin_filas = 18

for fila in range(inicio_filas, fin_filas + 1):
    fase = ws[f"C{fila}"].value
    if fase:
        nueva_fecha_inicio = st.date_input(f"Inicio de '{fase}'", value=date.today(), key=f"ini_{fila}")
        nueva_fecha_fin = st.date_input(f"Fin de '{fase}'", value=date.today(), key=f"fin_{fila}")
        ws[f"G{fila}"] = nueva_fecha_inicio
        ws[f"H{fila}"] = nueva_fecha_fin

# Actualizar celda con fecha de actualización
ws["E2"] = f"FECHA DE ACTUALIZACIÓN: {fecha_actualizacion.strftime('%d.%m.%y')}"
ws["B1"] = nombre_proyecto

# Guardar nuevo archivo
nombre_salida = f"cronograma_actualizado_{nombre_proyecto}.xlsx"
wb.save(nombre_salida)

# Botón para descarga
st.success("✅ Archivo actualizado correctamente.")
with open(nombre_salida, "rb") as file:
    st.download_button(
        "📥 Descargar archivo Excel actualizado",
        data=file.read(),
        file_name=nombre_salida,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
